"""Render a user's expenses to Excel (.xlsx) or PDF for download."""
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.timeutils import as_utc
from app.models.expense import Expense

_VIOLET = "7C3AED"
_INK = "13111C"
_MIST = "F6F4FF"
_LINE = "E7E3F5"


def _fmt_dt(dt: datetime) -> str:
    return as_utc(dt).strftime("%d %b %Y, %I:%M %p")


def _amount_int(value: Decimal) -> int:
    return int(Decimal(value))


def build_xlsx(expenses: list[Expense], total_amount: Decimal) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Name", "Category", "Amount (PKR)", "Note"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor=_VIOLET)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for e in expenses:
        ws.append([
            _fmt_dt(e.spent_at),
            e.name,
            e.category.name if e.category else "",
            _amount_int(e.amount),
            e.description or "",
        ])

    ws.append([])
    total_row_idx = ws.max_row + 1
    ws.append(["", "", "Total", _amount_int(total_amount), ""])
    ws.cell(row=total_row_idx, column=3).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=4).font = Font(bold=True)

    for width, col in zip([22, 36, 20, 14, 32], range(1, 6)):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(expenses: list[Expense], total_amount: Decimal, subtitle: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=16 * mm,
        bottomMargin=14 * mm,
        leftMargin=13 * mm,
        rightMargin=13 * mm,
        title="Spend Jot — Expenses",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "sj_title", parent=styles["Title"], textColor=colors.HexColor(f"#{_INK}"),
        fontSize=18, spaceAfter=2, alignment=0,
    )
    sub_style = ParagraphStyle(
        "sj_sub", parent=styles["Normal"], textColor=colors.HexColor("#6B6580"), fontSize=9,
    )
    cell = ParagraphStyle("sj_cell", parent=styles["Normal"], fontSize=8, leading=10)

    data: list[list] = [["Date", "Name", "Category", "Amount", "Note"]]
    for e in expenses:
        data.append([
            Paragraph(_fmt_dt(e.spent_at), cell),
            Paragraph(escape(e.name), cell),
            Paragraph(escape(e.category.name if e.category else ""), cell),
            Paragraph(f"₨ {_amount_int(e.amount):,}", cell),
            Paragraph(escape(e.description or ""), cell),
        ])
    data.append([
        "", "", Paragraph("<b>Total</b>", cell),
        Paragraph(f"<b>₨ {_amount_int(total_amount):,}</b>", cell), "",
    ])

    table = Table(data, colWidths=[30 * mm, 46 * mm, 27 * mm, 25 * mm, 36 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_VIOLET}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor(f"#{_MIST}")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEABOVE", (0, -1), (-1, -1), 0.6, colors.HexColor(f"#{_LINE}")),
    ]))

    doc.build([
        Paragraph("Spend Jot — Expenses", title_style),
        Paragraph(subtitle, sub_style),
        Spacer(1, 8),
        table,
    ])
    return buf.getvalue()
